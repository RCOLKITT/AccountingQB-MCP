#!/usr/bin/env python3
"""AccountingQB — local "Door 2" shim.

A localhost-only HTTP server that wraps the AccountingQB MCP connector so the app
can run as a downloadable desktop app (Tauri sidecar) or plain `python serve.py`,
rendered in the browser — without depending on the Cowork bridge. It:

  * binds 127.0.0.1 ONLY, with Host + Origin allowlist checks (DNS-rebind / CSRF safe),
  * invokes the connector's 131 tools IN-PROCESS (no stdio round-trip) via FastMCP's
    tool registry,
  * relays chat to Claude with the user's OWN Anthropic key (BYO key; nothing AI
    touches our servers),
  * connects QuickBooks via loopback OAuth (local Intuit app) or the hosted broker,
  * auto-creates ~/.accountingqb and picks a free port.

Ported from the sibling repo Hearth's `hearth-local/serve.js`; Python here because the
connector is Python/FastMCP. This is Phase 2a — the shim + wiring. The tabbed
Chat+Dashboard UI (2b) and the signed Tauri build (2c) build on top of it.

Run:  python accountingqb-local/serve.py         (opens http://127.0.0.1:4318)
Env:  ACCOUNTINGQB_PORT, ACCOUNTINGQB_DATA_DIR, ACCOUNTINGQB_NO_OPEN,
      ANTHROPIC_API_KEY, ANTHROPIC_MODEL, plus the connector's QB_* vars.
"""

from __future__ import annotations

import base64
import datetime
import hashlib
import inspect
import json
import os
import re
import secrets
import socket
import subprocess
import sys
import urllib.parse
from pathlib import Path

# --- resolve resources both in-repo and inside a PyInstaller bundle (Phase 2c) ---
# In a bundle, sys._MEIPASS is the extraction dir where the spec places manifest.json +
# artifact.html at the root and the accountingqb package (importable directly). In the
# repo, add mcpb/src to sys.path and read manifest.json from mcpb/.
_BUNDLE_DIR = getattr(sys, "_MEIPASS", None)
if _BUNDLE_DIR:
    _RES = Path(_BUNDLE_DIR)
    _MANIFEST_PATH = _RES / "manifest.json"
else:
    _REPO_ROOT = Path(__file__).resolve().parent.parent
    _RES = Path(__file__).resolve().parent
    _MANIFEST_PATH = _REPO_ROOT / "mcpb" / "manifest.json"
    sys.path.insert(0, str(_REPO_ROOT / "mcpb" / "src"))

import httpx  # noqa: E402
import uvicorn  # noqa: E402
from accountingqb import server as qb  # noqa: E402  (the connector: mcp, tools, tokens)
from accountingqb.context import get_ctx  # noqa: E402
from starlette.applications import Starlette  # noqa: E402
from starlette.middleware.base import BaseHTTPMiddleware  # noqa: E402
from starlette.requests import Request  # noqa: E402
from starlette.responses import (  # noqa: E402
    HTMLResponse,
    JSONResponse,
    PlainTextResponse,
    RedirectResponse,
    Response,
)
from starlette.routing import Route  # noqa: E402

# ---------------------------------------------------------------------------
# Config / local state (mirrors Hearth's ~/.hearth pattern)
# ---------------------------------------------------------------------------
DATA_DIR = Path(os.environ.get("ACCOUNTINGQB_DATA_DIR", Path.home() / ".accountingqb"))
DATA_DIR.mkdir(parents=True, exist_ok=True)
CONFIG_FILE = DATA_DIR / "config.json"
# Idempotency ledger for Coffer/Hearth-pushed expenses (key -> booking ref). A key
# already here is never re-booked (the Coffer contract requires idempotent success).
BOOKED_FILE = DATA_DIR / "coffer_booked.json"
# Saved Client Package report templates (per client): name/logo/period/comparison/sections/line-edits.
TEMPLATES_DIR = DATA_DIR / "templates"

# Cross-app pairing (identity-verified). The integration bridge is INERT until a
# pairing exists here — populated by the account-anchored link flow (same verified
# Google/email owns both apps) and required on every integration /mcp call. This is
# what prevents cross-user contamination: no pairing → no cross-app data, ever.
PAIRING_FILE = DATA_DIR / "pairing.json"
_IDENTITY_NS = (
    "aqb-coffer-link:v1:"  # both products hash the account email with this prefix
)


def identity_hash(email: str) -> str:
    """Deterministic, cross-product identity from a verified account email."""
    return hashlib.sha256(
        (_IDENTITY_NS + (email or "").strip().lower()).encode()
    ).hexdigest()


def _load_pairing() -> dict:
    try:
        return json.loads(PAIRING_FILE.read_text())
    except Exception:
        return {}


def _save_pairing(d: dict) -> dict:
    tmp = PAIRING_FILE.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(d, indent=2))
    tmp.replace(PAIRING_FILE)
    return d


# --- OAuth-style "Connect …" linking (redirect + PKCE) ---------------------------
# Where AccountingQB's own account lives (mint/status) and where the peer's authorize
# page lives. Both overridable for local end-to-end testing.
AQB_API_URL = os.environ.get("QB_API_URL", "https://accountingqb.com")
COFFER_API_URL = os.environ.get("COFFER_API_URL", "https://coffermoney.com")
LINK_STATE_FILE = (
    DATA_DIR / "link_state.json"
)  # in-flight PKCE verifier + state (single-use)


def _load_link_state() -> dict:
    try:
        return json.loads(LINK_STATE_FILE.read_text())
    except Exception:
        return {}


def _save_link_state(d: dict) -> dict:
    tmp = LINK_STATE_FILE.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(d, indent=2))
    tmp.replace(LINK_STATE_FILE)
    return d


def _pkce_pair() -> tuple[str, str]:
    """(verifier, S256 challenge) — the challenge is base64url(sha256(verifier)), unpadded."""
    verifier = secrets.token_urlsafe(64)
    challenge = (
        base64.urlsafe_b64encode(hashlib.sha256(verifier.encode()).digest())
        .rstrip(b"=")
        .decode()
    )
    return verifier, challenge


def _link_result_html(title: str, body: str, ok: bool = True) -> str:
    color = "#22d3ee" if ok else "#f87171"
    return (
        "<html><body style='font-family:system-ui;text-align:center;padding:64px;"
        "background:#0a0e1a;color:#e5e7eb'>"
        f"<h1 style='color:{color}'>{title}</h1><p>{body}</p>"
        "<script>setTimeout(()=>window.close(),2500)</script></body></html>"
    )


ALLOWED_HOSTS = {"127.0.0.1", "localhost"}
ALLOWED_ORIGIN_HOSTS = {"127.0.0.1", "localhost", "tauri.localhost"}

# BYO-key AI. Chat uses a mid-tier model by default; the user pays via their own key.
DEFAULT_MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-6")

# Intuit OAuth (local / BYO-Intuit-app path) — same endpoints as setup.py.
OAUTH_AUTH_URL = "https://appcenter.intuit.com/connect/oauth2"
OAUTH_TOKEN_URL = "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer"
OAUTH_SCOPES = "com.intuit.quickbooks.accounting"


def load_config() -> dict:
    try:
        return json.loads(CONFIG_FILE.read_text())
    except Exception:
        return {}


def save_config(patch: dict) -> dict:
    cfg = {**load_config(), **patch}
    tmp = CONFIG_FILE.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(cfg, indent=2))
    tmp.replace(CONFIG_FILE)
    return cfg


def _anthropic_key() -> str:
    return os.environ.get("ANTHROPIC_API_KEY") or load_config().get(
        "anthropic_api_key", ""
    )


def _manifest() -> dict:
    try:
        return json.loads(_MANIFEST_PATH.read_text())
    except Exception:
        return {}


def _server_version() -> str:
    return _manifest().get("version", "dev")


# Curated, READ-ONLY tool set exposed to the autonomous Chat loop. Writes are
# intentionally excluded — an autonomous agent must never mutate the books without a
# human in the loop (that stays in Door 1 / a future confirm-gated flow). Names are
# intersected with what actually exists + the manifest's readOnlyHint at runtime.
_CHAT_ALLOW = {
    "qb_company_info",
    "qb_list_companies",
    "qb_profit_loss",
    "qb_balance_sheet",
    "qb_cash_flow",
    "qb_monthly_burn_rate",
    "qb_runway_calculator",
    "qb_deduction_finder",
    "qb_anomaly_detection",
    "qb_find_duplicates",
    "qb_books_health_audit",
    "qb_tax_summary",
    "qb_schedule_c",
    "qb_t2125_summary",
    "qb_estimate_quarterly_tax",
    "qb_stripe_reconcile",
    "qb_list_transactions",
    "qb_search_transactions",
    "qb_trial_balance",
    "qb_uncategorized_transactions",
    "qb_1099_contractor_report",
    "qb_account_balance",
    "qb_tax_data_info",
}


def _anthropic_tools() -> list:
    """Anthropic tool defs (name/description/input_schema) for the read-only chat set."""
    registry = _tool_registry()
    out = []
    for t in _manifest().get("tools", []):
        name = t.get("name")
        if name in _CHAT_ALLOW and name in registry and t.get("readOnlyHint", False):
            out.append(
                {
                    "name": name,
                    "description": (t.get("description") or "")[:1024],
                    "input_schema": t.get("inputSchema")
                    or {"type": "object", "properties": {}},
                }
            )
    return out


_CHAT_SYSTEM = (
    "You are AccountingQB, a bookkeeping and tax-prep assistant connected to the user's real "
    "QuickBooks Online via local tools. Answer from the tools — never invent figures. Cite the "
    "numbers you used and keep answers concise. If no company is connected, say so and suggest "
    "connecting QuickBooks. You have READ-ONLY tools here; to change the books, the user should "
    "use AccountingQB in Claude. This is not tax or accounting advice."
)


# ---------------------------------------------------------------------------
# In-process tool invocation (FastMCP 1.29: tool.fn / tool.is_async)
# ---------------------------------------------------------------------------
def _tool_registry() -> dict:
    return qb.mcp._tool_manager._tools  # noqa: SLF001 — stable in mcp 1.29


async def call_tool(name: str, args: dict) -> object:
    tools = _tool_registry()
    tool = tools.get(name)
    if tool is None:
        raise KeyError(f"unknown tool: {name}")
    if getattr(tool, "context_kwarg", None):
        # A tool that wants FastMCP's request Context can't be called bare here;
        # none of the connector's tools need it today, but fail loud if that changes.
        raise RuntimeError(
            f"tool {name} requires a FastMCP Context (unsupported in the local shim)"
        )
    fn = tool.fn
    call_args = dict(args or {})
    # Additive-fields rule: peers may send fields a tool doesn't declare yet (e.g. Coffer passing
    # `type` to qb_list_accounts, whose param is `account_type`). Drop unknown kwargs instead of
    # raising a TypeError — unless the tool itself accepts **kwargs.
    try:
        params = inspect.signature(fn).parameters
        if not any(p.kind == inspect.Parameter.VAR_KEYWORD for p in params.values()):
            call_args = {k: v for k, v in call_args.items() if k in params}
    except (TypeError, ValueError):  # builtins / C-callables without a signature
        pass
    result = await fn(**call_args) if tool.is_async else fn(**call_args)
    return result


# ---------------------------------------------------------------------------
# Coffer/Hearth integration layer (the /mcp "integration dialect")
# ---------------------------------------------------------------------------
# For the contract's three tool names, /mcp returns STRUCTURED JSON (not the raw
# markdown the connector tools return), by reusing the connector's own vetted logic:
#   - qb_owner_draws            → {ytd, net, ...}  (reuse the audited markdown, parse the net)
#   - qb_estimate_quarterly_tax → {amount, due, period}  (reuse the vetted tax math; require
#                                  filing_status — NEVER guess a status → a wrong number)
#   - qb_record_owner_paid_expense → idempotent journal entry: DR review account / CR owner
#                                  equity (owner paid personally); confirm-gated.
# Reads are fail-closed: if the number can't be read unambiguously we return {error} and Coffer
# shows nothing — never a wrong figure (Constitution: always accurate).


def _load_booked() -> dict:
    try:
        return json.loads(BOOKED_FILE.read_text())
    except Exception:
        return {}


def _record_booked(key: str, ref: dict) -> None:
    b = _load_booked()
    b[key] = ref
    tmp = BOOKED_FILE.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(b, indent=2))
    tmp.replace(BOOKED_FILE)


def _parse_money(s: str):
    """Parse a fmt()/fmt_signed() amount from a line. Parens → negative. None if absent."""
    m = re.search(r"(\()?\$([\d,]+\.\d{2})(\))?", s)
    if not m:
        return None
    val = float(m.group(2).replace(",", ""))
    return -val if m.group(1) else val


async def _int_owner_draws(args: dict) -> dict:
    year = int(args.get("year") or datetime.date.today().year)
    md = str(await call_tool("qb_owner_draws", {"year": year}))
    net = None
    for line in md.splitlines():
        if "Net owner activity:" in line:
            net = _parse_money(line)
            break
    if net is None:
        if "No equity activity" in md:
            net = 0.0
        else:
            return {"error": "could not read owner draws", "summary": md}
    ytd = (
        round(-net, 2) if net < 0 else 0.0
    )  # money drawn OUT of the business = household income
    return {
        "year": year,
        "net": net,
        "ytd": ytd,
        "label": "net owner draws YTD (money taken out of the business)",
        "currency": "USD",
        "provenance": "qb_owner_draws",
        "summary": md,
    }


async def _int_estimate(args: dict) -> dict:
    fs = (args.get("filing_status") or "").strip()
    if not fs:
        # Never guess filing status — it materially changes the number.
        return {"error": "filing_status required", "needs": ["filing_status"]}
    ty = str(args.get("tax_year") or datetime.date.today().year)
    state = args.get("state") or ""
    md = str(
        await call_tool(
            "qb_estimate_quarterly_tax",
            {"tax_year": ty, "filing_status": fs, "state": state},
        )
    )
    if "no estimated payments are due" in md:  # net loss
        return {
            "amount": 0.0,
            "period": None,
            "due": None,
            "annual": 0.0,
            "note": "net loss — no estimated payments due",
            "provenance": "qb_estimate_quarterly_tax",
        }
    amount = annual = None
    for line in md.splitlines():
        if "Each quarterly payment:" in line and amount is None:
            amount = _parse_money(line)
        if "Total estimated annual tax:" in line and annual is None:
            annual = _parse_money(line)
    period = due = None
    for line in md.splitlines():
        st = line.strip()
        if st.startswith("Q") and "—" in st and ("Current" in st or "Upcoming" in st):
            m = re.match(r"Q(\d):\s*(.+?)\s*—", st)
            if m:
                period, due = f"Q{m.group(1)} {ty}", m.group(2).strip()
                break
    if amount is None or period is None:
        return {"error": "could not read estimate", "summary": md}
    return {
        "amount": amount,
        "due": due,
        "period": period,
        "annual": annual,
        "filing_status": fs,
        "state": state or "(auto)",
        "provenance": "qb_estimate_quarterly_tax",
    }


# A fresh QuickBooks company has no clearing account and often no owner-equity account. Reuse any
# existing owner-equity account before creating one, so we don't clutter the chart (Coffer never
# names accounts — choosing/creating them is our domain per the contract).
_OWNER_EQUITY_CANDIDATES = [
    "Owner's Equity",
    "Owners Equity",
    "Owner Equity",
    "Owner Investment",
    "Owner investments",
    "Owner's Investment",
    "Owner Contributions",
    "Member's Equity",
    "Opening Balance Equity",
]


def _parse_account_names(listing: str) -> dict:
    """Parse qb_list_accounts markdown ('### <Type>' headers + '- <name> (ID: <id>) | …') into
    {lowercased name: (name, account_type)}."""
    out: dict = {}
    cur = ""
    for line in str(listing or "").splitlines():
        line = line.strip()
        if line.startswith("### "):
            cur = line[4:].strip()
        elif line.startswith("- "):
            m = re.match(r"- (.+?) \(ID: ", line)
            if m:
                nm = m.group(1).strip()
                out[nm.lower()] = (nm, cur)
    return out


async def _ensure_account(
    name: str,
    account_type: str,
    sub_type: str,
    existing: dict,
    candidates: list | None = None,
) -> str | None:
    """Return the name of a usable account: the configured one if it exists, else an existing
    candidate (equity reuse), else create it on first use. None if creation fails."""
    if name.lower() in existing:
        return existing[name.lower()][0]
    for c in candidates or []:
        if c.lower() in existing:
            return existing[c.lower()][0]
    res = str(
        await call_tool(
            "qb_create_account",
            {"name": name, "account_type": account_type, "account_sub_type": sub_type},
        )
    )
    if "Created account" in res or any(
        w in res.lower() for w in ("duplicate", "already", "exists")
    ):
        return name  # created, or a concurrent create won the race
    return None


async def _int_owner_paid_expense(args: dict) -> dict:
    key = str(args.get("key") or "").strip()
    if not key:
        return {"ok": False, "error": "missing key"}
    if not args.get("confirmed"):
        return {"ok": False, "error": "confirmation required (confirmed:true)"}
    # Bind the write to the confirmed company: if Coffer names an expected realm and the
    # active QB company doesn't match, refuse — a company switch can't misroute a booking.
    expected_realm = str(args.get("expected_realm_id") or "").strip()
    if expected_realm:
        active_realm = str(getattr(get_ctx(), "realm_id", "") or "")
        if active_realm and active_realm != expected_realm:
            return {
                "ok": False,
                "error": "active QuickBooks company does not match expected_realm_id",
                "activeRealm": active_realm,
                "expectedRealm": expected_realm,
            }
    booked = _load_booked()
    if key in booked:  # idempotent: already booked → succeed again, never duplicate
        return {"ok": True, "key": key, "alreadyBooked": True, "qb": booked[key]}
    try:
        amt = round(abs(float(args.get("amount", 0))), 2)
    except (TypeError, ValueError):
        return {"ok": False, "error": "invalid amount"}
    if amt <= 0:
        return {"ok": False, "error": "amount must be non-zero"}
    date = str(args.get("date") or "")
    cfg = load_config()
    expense_acct = (
        args.get("expense_account")
        or cfg.get("owner_paid_expense_account")
        or "Owner-Paid Expenses (review)"
    )
    equity_acct = (
        args.get("equity_account")
        or cfg.get("owner_equity_account")
        or "Owner's Equity"
    )
    # First-use bootstrap: ensure both booking accounts exist (fresh QuickBooks companies have
    # neither the review clearing account nor an owner-equity account). Create the clearing account
    # if missing; reuse any existing owner-equity account before creating one.
    try:
        existing = _parse_account_names(str(await call_tool("qb_list_accounts", {})))
    except Exception:
        existing = {}
    resolved_expense = await _ensure_account(
        expense_acct, "Expense", "OfficeGeneralAdministrativeExpenses", existing
    )
    resolved_equity = await _ensure_account(
        equity_acct, "Equity", "OwnersEquity", existing, _OWNER_EQUITY_CANDIDATES
    )
    if not resolved_expense or not resolved_equity:
        return {
            "ok": False,
            "key": key,
            "error": "Could not ensure the booking accounts exist in QuickBooks.",
            "needs": f"an Expense clearing account ('{expense_acct}') and an owner Equity account",
        }
    expense_acct, equity_acct = resolved_expense, resolved_equity
    merchant = args.get("realMerchant") or args.get("merchant") or "Owner-paid"
    receipt = args.get("receipt") or {}
    memo = f"coffer:{key} | {merchant}"
    if args.get("note"):
        memo += f" | {args['note']}"
    if isinstance(receipt, dict) and receipt.get("orderId"):
        memo += f" | receipt {receipt['orderId']}"
    if args.get("category"):
        memo += f" | coffer-cat: {args['category']}"
    lines = [
        {
            "account_name": expense_acct,
            "amount": amt,
            "type": "Debit",
            "description": memo,
        },
        {
            "account_name": equity_acct,
            "amount": amt,
            "type": "Credit",
            "description": memo,
        },
    ]
    result = str(
        await call_tool(
            "qb_create_journal_entry",
            {"date": date, "lines_json": json.dumps(lines), "memo": memo},
        )
    )
    if "Journal entry created" in result:
        ref = {
            "account": expense_acct,
            "equity": equity_acct,
            "amount": amt,
            "date": date,
        }
        _record_booked(key, ref)
        return {
            "ok": True,
            "key": key,
            "booked": True,
            "amount": amt,
            "treatment": f"DR {expense_acct} / CR {equity_acct}",
            "message": "Booked as an owner-paid business expense (review account); credited owner's equity.",
        }
    return {
        "ok": False,
        "key": key,
        "error": result,
        "needs": f"QuickBooks accounts '{expense_acct}' (Expense) and '{equity_acct}' (Equity) must exist.",
    }


INTEGRATION_HANDLERS = {
    "qb_owner_draws": _int_owner_draws,
    "qb_estimate_quarterly_tax": _int_estimate,
    "qb_record_owner_paid_expense": _int_owner_paid_expense,
}


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
async def healthz(_req: Request) -> PlainTextResponse:
    return PlainTextResponse("ok")


async def api_status(_req: Request) -> JSONResponse:
    ctx = get_ctx()
    connected = bool(getattr(ctx, "refresh_token", "") or ctx.hosted_mode)
    return JSONResponse(
        {
            "version": _server_version(),
            "toolCount": len(_tool_registry()),
            "hostedMode": bool(ctx.hosted_mode),
            "connected": connected,
            "hasAnthropicKey": bool(_anthropic_key()),
            "realmId": getattr(ctx, "realm_id", "") or "",
        }
    )


async def api_clients(_req: Request) -> JSONResponse:
    """The connected QuickBooks companies (the bookkeeper's clients) + which one
    is active. Books data stays local — this is connection metadata only."""
    ctx = get_ctx()
    if not ctx.hosted_mode:
        # Self-hosted BYO-Intuit: one company per connection (the UI already
        # shows its name from qb_company_info).
        return JSONResponse(
            {
                "clients": (
                    [
                        {
                            "realmId": getattr(ctx, "realm_id", "") or "",
                            "companyName": "",
                            "active": True,
                        }
                    ]
                    if getattr(ctx, "refresh_token", "")
                    else []
                ),
                "multi": False,
            }
        )
    try:
        if not getattr(ctx, "hosted_loaded", False):
            qb._fetch_hosted_tokens(ctx)
    except Exception:
        pass
    active = getattr(ctx, "realm_id", "") or ""
    clients = [
        {
            "realmId": c.get("realmId", ""),
            "companyName": c.get("companyName") or c.get("realmId", ""),
            "active": c.get("realmId") == active,
        }
        for c in getattr(ctx, "hosted_companies", []) or []
    ]
    return JSONResponse({"clients": clients, "multi": len(clients) > 1})


async def api_clients_switch(req: Request) -> JSONResponse:
    """Switch the active client (company). Connection-level state, not a book
    write — but only ever driven by an explicit user click in the Clients view."""
    try:
        body = await req.json()
    except Exception:
        return JSONResponse({"error": "invalid JSON"}, status_code=400)
    realm = str((body or {}).get("realmId") or "").strip()
    if not realm:
        return JSONResponse({"error": "realmId required"}, status_code=400)
    try:
        result = await call_tool("qb_switch_company", {"realm_id": realm})
        ok = isinstance(result, str) and ("✅" in result or "Switched" in result)
        return JSONResponse({"ok": ok, "result": str(result)[:300]})
    except Exception as e:
        return JSONResponse({"error": f"{type(e).__name__}: {e}"}, status_code=500)


async def api_tools(_req: Request) -> JSONResponse:
    tools = _tool_registry()
    return JSONResponse(
        {
            "tools": [
                {"name": n, "description": (t.description or "").split("\n")[0]}
                for n, t in sorted(tools.items())
            ]
        }
    )


_WRITE_TOOLS_CACHE: set | None = None


def _is_write_tool(name: str) -> bool:
    """True if the tool mutates the books (readOnlyHint != true in the manifest). Any such tool
    reached via /mcp from the UI must carry confirmed:true — no silent writes (Constitution:
    write-safety, human-in-the-loop). The read-only chat loop never exposes these."""
    global _WRITE_TOOLS_CACHE
    if _WRITE_TOOLS_CACHE is None:
        _WRITE_TOOLS_CACHE = {
            t.get("name")
            for t in _manifest().get("tools", [])
            if not t.get("readOnlyHint")
        }
    return name in _WRITE_TOOLS_CACHE


async def mcp_call(req: Request) -> JSONResponse:
    try:
        body = await req.json()
    except Exception:
        return JSONResponse(
            {"isError": True, "error": "invalid JSON body"}, status_code=400
        )
    name = (body or {}).get("tool")
    args = (body or {}).get("args") or {}
    if not name:
        return JSONResponse(
            {"isError": True, "error": "missing 'tool'"}, status_code=400
        )
    # Integration dialect: the three Coffer-contract tools return structured JSON —
    # but only to a caller that PRESENTS the pairing secret (Coffer always does).
    # Dialect is selected by authentication: no secret presented → this is the app's
    # own UI (or a plain MCP client) and the call falls through to the normal tool
    # below, so the Tax tab etc. keep working whether or not a pairing exists.
    # A presented-but-wrong/stale secret is still refused (cross-user contamination
    # guard), and the write tool stays behind the confirmed:true gate either way.
    handler = INTEGRATION_HANDLERS.get(name)
    presented = req.headers.get("x-aqb-pairing") or (
        args.get("pairing_secret") if isinstance(args, dict) else None
    )
    if handler is not None and presented is not None:
        pairing = _load_pairing()
        if not pairing.get("pairing_secret"):
            return JSONResponse(
                {
                    "error": "AccountingQB isn't linked to a Coffer account yet — pair the two apps first.",
                    "needs": "pairing",
                },
                status_code=403,
            )
        if not secrets.compare_digest(str(presented), str(pairing["pairing_secret"])):
            return JSONResponse(
                {"error": "invalid or missing pairing secret"}, status_code=403
            )
        if isinstance(args, dict):
            args.pop("pairing_secret", None)  # don't pass the secret through to tools
        try:
            return JSONResponse(await handler(args))
        except Exception as e:
            return JSONResponse({"error": f"{type(e).__name__}: {e}"})
    # No silent writes: any book-mutating tool must be explicitly confirmed by the user. The UI
    # gathers the fields in a confirm card and re-sends with confirmed:true (see artifact.html).
    if _is_write_tool(name) and not (isinstance(args, dict) and args.get("confirmed")):
        return JSONResponse(
            {
                "needsConfirm": True,
                "tool": name,
                "error": f"{name} changes your books — confirm before it runs.",
            }
        )
    if isinstance(args, dict):
        args.pop("confirmed", None)  # UI flag; never forwarded to the tool
    try:
        result = await call_tool(name, args)
        return JSONResponse({"ok": True, "result": result})
    except KeyError as e:
        return JSONResponse({"isError": True, "error": str(e)}, status_code=404)
    except Exception as e:  # tool-level errors surface as friendly messages, not 500s
        return JSONResponse({"isError": True, "error": f"{type(e).__name__}: {e}"})


async def sample(req: Request) -> JSONResponse:
    """BYO-key Claude relay. NO proxy — the key never leaves the machine."""
    key = _anthropic_key()
    if not key:
        return JSONResponse(
            {
                "needsKey": True,
                "error": "No Anthropic API key set. Add one to enable chat.",
            }
        )
    try:
        body = await req.json()
    except Exception:
        body = {}
    system = str(body.get("system") or "")
    messages = body.get("messages")
    if not messages:
        # accept a simple {system, ctx} shape too
        messages = [
            {
                "role": "user",
                "content": json.dumps(body.get("ctx") or body.get("prompt") or {}),
            }
        ]
    model = body.get("model") or DEFAULT_MODEL
    try:
        async with httpx.AsyncClient(timeout=120) as client:
            r = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "content-type": "application/json",
                    "x-api-key": key,
                    "anthropic-version": "2023-06-01",
                },
                json={
                    "model": model,
                    "max_tokens": 2048,
                    "system": system,
                    "messages": messages,
                },
            )
        if r.status_code != 200:
            return JSONResponse(
                {"error": f"Anthropic {r.status_code}: {r.text[:400]}"}, status_code=502
            )
        data = r.json()
        text = "".join(part.get("text", "") for part in data.get("content", []))
        return JSONResponse({"text": text})
    except Exception as e:
        return JSONResponse({"error": f"{type(e).__name__}: {e}"}, status_code=502)


async def chat(req: Request) -> JSONResponse:
    """Agentic chat loop: Claude (BYO key) picks read-only tools, we run them via the
    in-process registry and feed results back, until it answers. Returns the final text
    plus a trace of the tools it called. Bounded to keep cost sane on the user's key."""
    key = _anthropic_key()
    if not key:
        return JSONResponse(
            {"needsKey": True, "error": "Add your Anthropic API key to use chat."}
        )
    try:
        body = await req.json()
    except Exception:
        body = {}
    messages = body.get("messages") or []
    if not messages:
        return JSONResponse({"error": "no messages"}, status_code=400)
    model = body.get("model") or DEFAULT_MODEL
    tools = _anthropic_tools()
    trace: list = []
    headers = {
        "content-type": "application/json",
        "x-api-key": key,
        "anthropic-version": "2023-06-01",
    }
    try:
        async with httpx.AsyncClient(timeout=180) as client:
            for _round in range(6):  # bound tool-use rounds
                r = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers=headers,
                    json={
                        "model": model,
                        "max_tokens": 2048,
                        "system": _CHAT_SYSTEM,
                        "tools": tools,
                        "messages": messages,
                    },
                )
                if r.status_code != 200:
                    return JSONResponse(
                        {"error": f"Anthropic {r.status_code}: {r.text[:400]}"},
                        status_code=502,
                    )
                data = r.json()
                content = data.get("content", [])
                messages.append({"role": "assistant", "content": content})
                tool_uses = [b for b in content if b.get("type") == "tool_use"]
                if data.get("stop_reason") == "tool_use" and tool_uses:
                    results = []
                    for tu in tool_uses:
                        trace.append({"tool": tu.get("name"), "args": tu.get("input")})
                        try:
                            out = await call_tool(tu["name"], tu.get("input") or {})
                            results.append(
                                {
                                    "type": "tool_result",
                                    "tool_use_id": tu["id"],
                                    "content": str(out)[:20000],
                                }
                            )
                        except Exception as e:
                            results.append(
                                {
                                    "type": "tool_result",
                                    "tool_use_id": tu["id"],
                                    "content": f"error: {e}",
                                    "is_error": True,
                                }
                            )
                    messages.append({"role": "user", "content": results})
                    continue
                text = "".join(
                    b.get("text", "") for b in content if b.get("type") == "text"
                )
                return JSONResponse({"reply": text, "trace": trace})
        return JSONResponse(
            {
                "reply": "I ran several tools but couldn't finish — try narrowing the question.",
                "trace": trace,
            }
        )
    except Exception as e:
        return JSONResponse({"error": f"{type(e).__name__}: {e}"}, status_code=502)


async def api_config(req: Request) -> JSONResponse:
    """Persist local settings from the UI (Anthropic key, optional Intuit app creds)."""
    try:
        body = await req.json()
    except Exception:
        body = {}
    patch = {}
    for k_in, k_cfg in (
        ("anthropicApiKey", "anthropic_api_key"),
        ("qbClientId", "qb_client_id"),
        ("qbClientSecret", "qb_client_secret"),
        ("licenseKey", "license_key"),
    ):
        if body.get(k_in):
            patch[k_cfg] = str(body[k_in])
    if patch:
        save_config(patch)
    # Saving a license should immediately reuse the existing profile's hosted QuickBooks company.
    if patch.get("license_key"):
        _bootstrap_profile()
    return JSONResponse({"ok": True, "hasAnthropicKey": bool(_anthropic_key())})


async def whoami(_req: Request) -> JSONResponse:
    """Peer-identity probe. Coffer calls this to confirm it's really talking to AccountingQB
    and to read the active QB realm (to pass back as expected_realm_id). No secret exposed.
    """
    p = _load_pairing()
    ctx = get_ctx()
    return JSONResponse(
        {
            "app": "accountingqb",
            "version": _server_version(),
            "paired": bool(p.get("pairing_secret")),
            "peerProduct": p.get("peer_product"),
            "realm": str(getattr(ctx, "realm_id", "") or ""),
        }
    )


async def pair(req: Request) -> JSONResponse:
    """Establish the cross-app pairing on this shim. The pairingSecret comes from the
    identity-verified account link flow (same verified email owns both apps); the AccountingQB
    desktop app fetches it for its account and posts it here. Localhost-only (same trust model
    as the rest of the shim)."""
    try:
        body = await req.json()
    except Exception:
        body = {}
    secret = body.get("pairingSecret") or body.get("pairing_secret")
    if not secret:
        return JSONResponse(
            {"ok": False, "error": "pairingSecret required"}, status_code=400
        )
    rec = {
        "pairing_secret": str(secret),
        "peer_product": body.get("peerProduct") or "coffer",
        "peer_identity": body.get("peerIdentity") or "",
        "aqb_identity": body.get("aqbIdentity") or "",
        "created_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    }
    _save_pairing(rec)
    return JSONResponse(
        {"ok": True, "paired": True, "peerProduct": rec["peer_product"]}
    )


async def unpair(_req: Request) -> JSONResponse:
    _save_pairing({})
    return JSONResponse({"ok": True, "paired": False})


# --- OAuth-style "Connect Coffer" (AccountingQB initiates) -----------------------
async def link_connect(req: Request) -> RedirectResponse:
    """Start linking the peer app from AccountingQB. Generates a PKCE verifier + state, then sends
    the browser to the peer's authorize page with a loopback redirect back to THIS shim. The user
    logs into the peer, consents, and is bounced to /link/callback. Symmetric to Coffer's own
    'Connect AccountingQB' button — either app can start the link."""
    peer = req.query_params.get("peer", "coffer")
    verifier, challenge = _pkce_pair()
    state = secrets.token_urlsafe(16)
    _save_link_state(
        {
            "verifier": verifier,
            "state": state,
            "peer": peer,
            "ts": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        }
    )
    port = req.url.port or PORT
    redirect_uri = f"http://127.0.0.1:{port}/link/callback"
    base = COFFER_API_URL.rstrip("/")  # only 'coffer' peer today
    params = urllib.parse.urlencode(
        {
            "peer": "accountingqb",
            "redirect_uri": redirect_uri,
            "state": state,
            "code_challenge": challenge,
            "code_challenge_method": "S256",
        }
    )
    return RedirectResponse(f"{base}/link/authorize?{params}", status_code=302)


async def link_callback(req: Request) -> HTMLResponse:
    """Return leg of an AccountingQB-initiated link. Verify state, redeem the code at the peer with
    our PKCE verifier, and store the resulting pairing secret. One shared secret gates calls in both
    directions — storing it here means Coffer↔AccountingQB is live on this machine."""
    p = req.query_params
    st = _load_link_state()
    if p.get("error"):
        _save_link_state({})
        return HTMLResponse(
            _link_result_html(
                "Link canceled",
                "You can close this tab and try again from AccountingQB.",
                ok=False,
            ),
            200,
        )
    code = p.get("code")
    state = p.get("state")
    if (
        not code
        or not state
        or not st.get("state")
        or not secrets.compare_digest(state, str(st.get("state")))
    ):
        return HTMLResponse(
            _link_result_html(
                "Link failed", "State mismatch — please retry the connection.", ok=False
            ),
            400,
        )
    peer = st.get("peer", "coffer")
    base = COFFER_API_URL.rstrip("/")
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                f"{base}/api/link/redeem",
                json={"code": code, "codeVerifier": st.get("verifier", "")},
            )
        data = r.json() if r.status_code == 200 else {}
    except Exception:
        return HTMLResponse(
            _link_result_html(
                "Link failed", f"Could not reach {peer}. Is it online?", ok=False
            ),
            502,
        )
    secret = data.get("pairingSecret")
    if not secret:
        detail = data.get("error") or "the peer rejected the code"
        return HTMLResponse(
            _link_result_html("Link failed", f"{detail}. Please retry.", ok=False), 400
        )
    rec = _load_pairing()
    rec.update(
        {
            "pairing_secret": str(secret),
            "peer_product": data.get("peerProduct") or peer,
            "peer_base_url": base,
            "linked_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        }
    )
    _save_pairing(rec)
    _save_link_state({})
    return HTMLResponse(
        _link_result_html(
            f"{peer.title()} connected ✓",
            "You can close this tab and return to AccountingQB.",
        ),
        200,
    )


async def link_refresh(_req: Request) -> JSONResponse:
    """Pull this account's pairing secret from the web and store it locally. Used after a
    peer-initiated 'Connect AccountingQB' link (the secret was minted when the peer redeemed our
    code) so incoming Coffer→AccountingQB calls are accepted. Idempotent; needs a license key.
    """
    lic = getattr(qb, "LICENSE_KEY", "") or load_config().get("license_key", "")
    if not lic:
        return JSONResponse(
            {"paired": False, "error": "no license key configured"}, status_code=400
        )
    base = AQB_API_URL.rstrip("/")
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            r = await client.get(f"{base}/api/link/status", params={"key": lic})
        data = r.json() if r.status_code == 200 else {}
    except Exception as e:
        return JSONResponse(
            {"paired": False, "error": f"{type(e).__name__}"}, status_code=502
        )
    if data.get("paired") and data.get("pairingSecret"):
        rec = _load_pairing()
        rec.update(
            {
                "pairing_secret": str(data["pairingSecret"]),
                "peer_product": data.get("peerProduct") or "coffer",
                "peer_identity": data.get("peerIdentity")
                or rec.get("peer_identity", ""),
                "peer_base_url": rec.get("peer_base_url") or COFFER_API_URL.rstrip("/"),
                "linked_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            }
        )
        _save_pairing(rec)
        return JSONResponse({"paired": True, "peerProduct": rec["peer_product"]})
    return JSONResponse({"paired": False})


def _redirect_uri(req: Request) -> str:
    # Loopback redirect back to THIS shim. Local (BYO Intuit app) users must register
    # this exact URI in their Intuit app; the hosted broker path avoids this entirely.
    return f"http://localhost:{req.url.port or PORT}/callback"


async def oauth_start(req: Request) -> RedirectResponse:
    cfg = load_config()
    client_id = os.environ.get("QB_CLIENT_ID") or cfg.get("qb_client_id")
    if not client_id:
        # Hosted broker path (no Intuit app needed): hand off to the web connect flow.
        lic = getattr(qb, "LICENSE_KEY", "") or cfg.get("license_key", "")
        api = os.environ.get("QB_API_URL", "https://accountingqb.com").rstrip("/")
        dest = f"{api}/setup-wizard" + (
            f"?key={urllib.parse.quote(lic)}" if lic else ""
        )
        return RedirectResponse(dest, status_code=302)
    state = secrets.token_urlsafe(16)
    save_config({"oauth_state": state})
    params = urllib.parse.urlencode(
        {
            "client_id": client_id,
            "response_type": "code",
            "scope": OAUTH_SCOPES,
            "redirect_uri": _redirect_uri(req),
            "state": state,
        }
    )
    return RedirectResponse(f"{OAUTH_AUTH_URL}?{params}", status_code=302)


async def oauth_callback(req: Request) -> HTMLResponse:
    p = req.query_params
    if p.get("error"):
        return HTMLResponse(
            f"<h1>Authorization failed</h1><p>{p.get('error_description', p.get('error'))}</p>",
            400,
        )
    code = p.get("code")
    realm = p.get("realmId")
    if not code:
        return HTMLResponse("<h1>Missing authorization code</h1>", 400)
    if p.get("state") and p.get("state") != load_config().get("oauth_state"):
        return HTMLResponse(
            "<h1>State mismatch</h1><p>Please retry the connection.</p>", 400
        )
    cfg = load_config()
    client_id = os.environ.get("QB_CLIENT_ID") or cfg.get("qb_client_id")
    client_secret = os.environ.get("QB_CLIENT_SECRET") or cfg.get("qb_client_secret")
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                OAUTH_TOKEN_URL,
                data={
                    "grant_type": "authorization_code",
                    "code": code,
                    "redirect_uri": _redirect_uri(req),
                },
                auth=(client_id, client_secret),
                headers={"Accept": "application/json"},
            )
        if r.status_code != 200:
            return HTMLResponse(
                f"<h1>Token exchange failed</h1><pre>{r.text[:500]}</pre>", 400
            )
        tok = r.json()
    except Exception as e:
        return HTMLResponse(f"<h1>Token exchange error</h1><pre>{e}</pre>", 502)

    refresh = tok.get("refresh_token", "")
    if refresh:
        qb._save_token(refresh)  # encrypted to ~/.accountingqb via the connector
        ctx = get_ctx()
        ctx.refresh_token = refresh
        if realm:
            ctx.realm_id = realm
            save_config({"realm_id": realm})
    return HTMLResponse(
        "<html><body style='font-family:system-ui;text-align:center;padding:64px;background:#0a0e1a;color:#e5e7eb'>"
        "<h1 style='color:#22d3ee'>&#10003; QuickBooks connected</h1>"
        "<p>You can close this tab and return to AccountingQB.</p>"
        "<script>setTimeout(()=>window.close(),2500)</script></body></html>"
    )


_ARTIFACT_PATH = _RES / "artifact.html"
_VENDOR_DIR = _RES / "vendor"  # bundled JS (pdfmake) shipped next to the artifact
# CHANGELOG.md (repo root; bundled into the sidecar) powers the in-app "What's new" panel.
_CHANGELOG_PATH = (
    (_RES / "CHANGELOG.md") if _BUNDLE_DIR else (_REPO_ROOT / "CHANGELOG.md")
)


async def vendor(req: Request) -> Response:
    """Serve vendored front-end libraries (pdfmake) from disk / the PyInstaller bundle. Localhost-only
    (LocalOnly middleware); no path traversal (name is a bare filename)."""
    name = req.path_params.get("name", "")
    if not re.fullmatch(r"[A-Za-z0-9_.-]+", name):
        return PlainTextResponse("bad request", status_code=400)
    f = _VENDOR_DIR / name
    if not f.exists() or not f.is_file():
        return PlainTextResponse("not found", status_code=404)
    ct = (
        "application/javascript" if name.endswith(".js") else "application/octet-stream"
    )
    return Response(
        f.read_bytes(), media_type=ct, headers={"cache-control": "max-age=86400"}
    )


# --- Client Package: real multi-sheet, formatted .xlsx (openpyxl) from already-gathered sections ---
_XLSX_NUM = re.compile(r"^\**[-–]?\(?[-–]?\$?[\d,]+\.?\d*\)?%?\**$")
_XLSX_FMT = "#,##0.00;(#,##0.00)"


def _xlsx_cell(s):
    """Return (value, is_number). Parse '$1,234.00' / '(20,224)' into real numbers so the workbook is
    sortable/sum-able; leave non-numeric strings as text. Never re-computes — just types the value.
    """
    s = str(s or "").strip()
    if not s:
        return "", False
    if (
        _XLSX_NUM.match(s)
        and any(c.isdigit() for c in s)
        and not s.strip("*").endswith("%")
    ):
        neg = s.lstrip("*").startswith(("(", "-", "–"))
        try:
            return (-1 if neg else 1) * float(re.sub(r"[^\d.]", "", s)), True
        except ValueError:
            return s, False
    return s, False


def _xlsx_sheet_title(t, used):
    t = re.sub(r"[\\/*?:\[\]]", " ", str(t))[:31].strip() or "Sheet"
    base, i = t, 2
    while t in used:
        t = (base[:27] + f" {i}")[:31]
        i += 1
    used.add(t)
    return t


async def export_xlsx(req: Request) -> Response:
    """Build a formatted, multi-sheet Excel workbook from the sections the UI already gathered
    (each figure is a live QuickBooks value — the server only formats, never computes).
    """
    try:
        body = await req.json()
    except Exception:
        return JSONResponse({"error": "bad json"}, status_code=400)
    try:
        import io

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except Exception:
        return JSONResponse(
            {"error": "Excel export unavailable (openpyxl not installed)."},
            status_code=501,
        )

    client = str(body.get("client") or "Client")
    period = body.get("period") or {}
    narrative = str(body.get("narrative") or "")
    sections = body.get("sections") or []
    bold = Font(bold=True)

    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "AccountingQB — Financial Package"
    cover["A1"].font = Font(bold=True, size=16)
    cover["A3"], cover["B3"] = "Client", client
    cover["A4"], cover["B4"] = (
        "Period",
        f"{period.get('start', '')} to {period.get('end', '')}",
    )
    cover["A3"].font = cover["A4"].font = bold
    if narrative:
        cover["A6"] = "Management commentary"
        cover["A6"].font = bold
        cover["A7"] = narrative
        cover["A7"].alignment = Alignment(wrap_text=True, vertical="top")
        cover.merge_cells("A7:F22")
    cover.column_dimensions["A"].width = 22
    cover.column_dimensions["B"].width = 44

    used = {"Cover"}
    for sec in sections:
        ws = wb.create_sheet(_xlsx_sheet_title(sec.get("title") or "Report", used))
        parsed = sec.get("parsed") or {}
        if parsed.get("kind") == "table":
            header = parsed.get("header") or []
            ws.append(header)
            for c in ws[1]:
                c.font = bold
            for r in parsed.get("rows") or []:
                vals, isnum = [], []
                for cell in r:
                    v, n = _xlsx_cell(cell)
                    vals.append(v)
                    isnum.append(n)
                ws.append(vals)
                total = (
                    str(r[0] if r else "")
                    .lower()
                    .lstrip("*")
                    .startswith(("total", "net ", "gross ", "subtotal"))
                )
                for i, cell in enumerate(ws[ws.max_row]):
                    if i < len(isnum) and isnum[i]:
                        cell.number_format = _XLSX_FMT
                        cell.alignment = Alignment(horizontal="right")
                    if total:
                        cell.font = bold
            ws.freeze_panes = "A2"
            for i, _ in enumerate(header, 1):
                ws.column_dimensions[get_column_letter(i)].width = 32 if i == 1 else 16
        else:
            cmp_label = parsed.get("compareLabel")
            if cmp_label:
                ws.append(["", "Current", cmp_label])
                for c in ws[ws.max_row]:
                    c.font = bold
                ws.freeze_panes = "A2"
            for it in parsed.get("items") or []:
                if it.get("sub"):
                    ws.append([it["sub"]])
                    ws[ws.max_row][0].font = Font(bold=True, color="FF0A5C39")
                    continue
                v, n = _xlsx_cell(it.get("val"))
                if cmp_label:
                    v2, n2 = _xlsx_cell(it.get("val2"))
                    ws.append([it.get("label", ""), v, v2])
                else:
                    ws.append([it.get("label", ""), v])
                    n2 = False
                row = ws[ws.max_row]
                if it.get("total"):
                    for cell in row:
                        cell.font = bold
                if n:
                    row[1].number_format = _XLSX_FMT
                    row[1].alignment = Alignment(horizontal="right")
                if cmp_label and n2:
                    row[2].number_format = _XLSX_FMT
                    row[2].alignment = Alignment(horizontal="right")
            ws.column_dimensions["A"].width = 46
            ws.column_dimensions["B"].width = 18
            if cmp_label:
                ws.column_dimensions["C"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    fname = re.sub(r"[^\w-]+", "_", client) + "_Package.xlsx"
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"content-disposition": f'attachment; filename="{fname}"'},
    )


# --- Client Package templates (reusable per-client report config; local files only) ---
def _tmpl_safe(name: str):
    s = re.sub(r"[^A-Za-z0-9 _.-]", "", str(name or "")).strip()[:60]
    return s or None


async def templates_list_or_save(req: Request) -> JSONResponse:
    if req.method == "POST":
        try:
            body = await req.json()
        except Exception:
            body = {}
        name = _tmpl_safe(body.get("name"))
        if not name:
            return JSONResponse({"error": "template name required"}, status_code=400)
        TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
        tmp = TEMPLATES_DIR / (name + ".json.tmp")
        tmp.write_text(
            json.dumps({"name": name, "config": body.get("config") or {}}, indent=2)
        )
        tmp.replace(TEMPLATES_DIR / (name + ".json"))
        return JSONResponse({"ok": True, "name": name})
    names = (
        sorted(f.stem for f in TEMPLATES_DIR.glob("*.json"))
        if TEMPLATES_DIR.exists()
        else []
    )
    return JSONResponse({"templates": names})


async def template_get_or_delete(req: Request) -> JSONResponse:
    name = _tmpl_safe(req.path_params.get("name"))
    if not name:
        return JSONResponse({"error": "bad name"}, status_code=400)
    f = TEMPLATES_DIR / (name + ".json")
    if req.method == "DELETE":
        try:
            f.unlink()
        except FileNotFoundError:
            pass
        return JSONResponse({"ok": True})
    if not f.exists():
        return JSONResponse({"error": "not found"}, status_code=404)
    try:
        return JSONResponse(json.loads(f.read_text()))
    except Exception:
        return JSONResponse({"error": "corrupt template"}, status_code=500)


def _parse_changelog(text: str) -> "list[dict]":
    """Split CHANGELOG.md into [{version, date, notes}] by `## <version> — <date>` headings.
    Order preserved (newest first, as written). The version token is the first word after `##`,
    with surrounding brackets stripped so both `## 0.2.0` and `## [0.2.0]` parse."""
    entries: list[dict] = []
    cur: dict | None = None
    for line in text.splitlines():
        if line.startswith("## "):
            head = line[3:].strip()
            tok = head.split()[0].strip("[]") if head.split() else ""
            date = ""
            if "—" in head:
                date = head.split("—", 1)[1].strip()
            elif " - " in head:
                date = head.split(" - ", 1)[1].strip()
            cur = {"version": tok, "date": date, "notes": ""}
            entries.append(cur)
        elif cur is not None:
            cur["notes"] += line + "\n"
    for e in entries:
        e["notes"] = e["notes"].strip()
    return entries


async def api_whatsnew(req: Request) -> JSONResponse:
    """The release-notes entry for a version (defaults to the running app version), for the
    in-app "What's new" panel. Read-only, localhost-only. Returns {} if there's no changelog.
    """
    try:
        entries = _parse_changelog(_CHANGELOG_PATH.read_text())
    except Exception:
        return JSONResponse({})
    if not entries:
        return JSONResponse({})
    want = req.query_params.get("v") or os.environ.get("ACCOUNTINGQB_APP_VERSION", "")
    match = next((e for e in entries if e["version"] == want), None) or entries[0]
    return JSONResponse(match)


async def index(_req: Request) -> HTMLResponse:
    # Phase 2b: the tabbed Chat + Dashboard artifact. Loaded from disk in dev; the
    # Tauri build (2c) will embed it. Falls back to the Phase-2a status shell.
    try:
        return HTMLResponse(_ARTIFACT_PATH.read_text())
    except Exception:
        return HTMLResponse(_INDEX_HTML)


# ---------------------------------------------------------------------------
# Security: localhost-only Host + Origin allowlist (port Hearth serve.js:487-500)
# ---------------------------------------------------------------------------
class LocalOnly(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        host = (request.headers.get("host") or "").split(":")[0]
        if host and host not in ALLOWED_HOSTS:
            return PlainTextResponse("forbidden host", status_code=403)
        origin = request.headers.get("origin")
        if origin:
            oh = urllib.parse.urlparse(origin).hostname
            if oh not in ALLOWED_ORIGIN_HOSTS:
                return PlainTextResponse("forbidden origin", status_code=403)
        return await call_next(request)


routes = [
    Route("/", index),
    Route("/vendor/{name}", vendor),
    Route("/export/xlsx", export_xlsx, methods=["POST"]),
    Route("/templates", templates_list_or_save, methods=["GET", "POST"]),
    Route("/templates/{name}", template_get_or_delete, methods=["GET", "DELETE"]),
    Route("/healthz", healthz),
    Route("/api/status", api_status),
    Route("/api/whatsnew", api_whatsnew),
    Route("/api/clients", api_clients),
    Route("/api/clients/switch", api_clients_switch, methods=["POST"]),
    Route("/api/tools", api_tools),
    Route("/api/config", api_config, methods=["POST"]),
    Route("/whoami", whoami),
    Route("/pair", pair, methods=["POST"]),
    Route("/unpair", unpair, methods=["POST"]),
    Route("/link/connect", link_connect),
    Route("/link/callback", link_callback),
    Route("/link/refresh", link_refresh, methods=["POST"]),
    Route("/mcp", mcp_call, methods=["POST"]),
    Route("/sample", sample, methods=["POST"]),
    Route("/chat", chat, methods=["POST"]),
    Route("/oauth/start", oauth_start),
    Route("/callback", oauth_callback),
]

app = Starlette(routes=routes, middleware=[])
app.add_middleware(LocalOnly)


# ---------------------------------------------------------------------------
# Boot: free port + auto-open browser (port Hearth main.rs / serve.js)
# ---------------------------------------------------------------------------
def pick_port() -> int:
    env = os.environ.get("ACCOUNTINGQB_PORT") or os.environ.get("PORT")
    if env:
        return int(env)
    for p in (
        4318,
        4319,
        4320,
    ):  # 4318 = the port Coffer/Hearth's contract expects (Hearth is 4317)
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            if s.connect_ex(("127.0.0.1", p)) != 0:
                return p
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return s.getsockname()[1]


PORT = pick_port()

_INDEX_HTML = """<!doctype html><html><head><meta charset=utf-8>
<title>AccountingQB</title><meta name=viewport content="width=device-width,initial-scale=1">
<style>body{font-family:system-ui;background:#0a0e1a;color:#e5e7eb;max-width:720px;margin:0 auto;padding:40px}
a.btn,button{background:#fff;color:#0a0e1a;border:0;border-radius:10px;padding:10px 16px;font-weight:600;cursor:pointer;text-decoration:none;display:inline-block}
.muted{color:#94a3b8}.card{border:1px solid rgba(255,255,255,.1);border-radius:14px;padding:20px;margin:16px 0;background:#131a2e}
input{width:100%;padding:10px;border-radius:8px;border:1px solid rgba(255,255,255,.15);background:#0d1220;color:#fff;margin:6px 0}
pre{white-space:pre-wrap;background:#0d1220;padding:12px;border-radius:8px;max-height:320px;overflow:auto}</style></head>
<body>
<h1><span style="color:#22d3ee">Accounting</span><span style="color:#60a5fa">QB</span> <span class=muted style="font-size:14px">local</span></h1>
<p class=muted id=status>Loading status…</p>
<div class=card><h3>1 · Connect QuickBooks</h3>
<p class=muted>Opens the QuickBooks consent flow. No data leaves your machine.</p>
<a class=btn href="/oauth/start">Connect QuickBooks</a></div>
<div class=card><h3>2 · Add your Anthropic key <span class=muted>(for chat)</span></h3>
<p class=muted>Stored only in ~/.accountingqb. Chat calls go from your machine straight to Anthropic.</p>
<input id=key type=password placeholder="sk-ant-…"><button onclick=saveKey()>Save key</button></div>
<div class=card><h3>Probe a tool</h3>
<button onclick="run('qb_server_info')">qb_server_info</button>
<button onclick="run('qb_tax_data_info')">qb_tax_data_info</button>
<pre id=out class=muted>—</pre></div>
<p class=muted style="font-size:12px">Phase 2a shell. The tabbed Chat + Dashboard UI lands in Phase 2b.</p>
<script>
async function refresh(){const s=await (await fetch('/api/status')).json();
document.getElementById('status').textContent=`v${s.version} · ${s.toolCount} tools · QuickBooks ${s.connected?'connected':'not connected'} · Anthropic key ${s.hasAnthropicKey?'set':'missing'}`;}
async function saveKey(){const k=document.getElementById('key').value.trim();if(!k)return;
await fetch('/api/config',{method:'POST',headers:{'content-type':'application/json'},body:JSON.stringify({anthropicApiKey:k})});
document.getElementById('key').value='';refresh();}
async function run(tool){const o=document.getElementById('out');o.textContent='Running '+tool+'…';
const r=await fetch('/mcp',{method:'POST',headers:{'content-type':'application/json'},body:JSON.stringify({tool,args:{}})});
const j=await r.json();o.textContent=typeof j.result==='string'?j.result:JSON.stringify(j,null,2);}
refresh();
</script></body></html>"""


def _bootstrap_profile() -> None:
    """Reuse the saved account profile: hand the license to the connector and pre-load the user's
    already-connected QuickBooks company (hosted mode) so the desktop is configured on launch — the
    SAME account as the Cowork plugin, no re-OAuth, no duplicate. Non-fatal if offline/unlicensed.
    """
    lic = os.environ.get("QB_LICENSE_KEY") or load_config().get("license_key", "")
    if not lic:
        return
    try:
        qb.LICENSE_KEY = lic
        ctx = get_ctx()
        ctx.license_key = lic
        if qb._fetch_hosted_tokens(ctx):
            print(
                f"  Profile loaded (hosted): QuickBooks company realm {getattr(ctx, 'realm_id', '') or '—'}"
            )
        else:
            print("  License set; no connected QuickBooks company found for it yet.")
    except Exception as e:  # pragma: no cover - network/offline
        print(f"  (profile bootstrap skipped: {type(e).__name__})")


def _bootstrap_pairing() -> None:
    """Restore the Coffer pairing from the web on launch. The web (account_links, keyed by license)
    is the source of truth; the local pairing.json is only a cache that a restart / fresh install /
    unpair can empty — which left whoami reporting paired:false and Coffer firing into an inert peer.
    Re-pulling here guarantees the shim comes up paired whenever the account is linked. Non-fatal.
    """
    lic = os.environ.get("QB_LICENSE_KEY") or load_config().get("license_key", "")
    if not lic:
        return
    base = AQB_API_URL.rstrip("/")
    try:
        with httpx.Client(timeout=15) as client:
            r = client.get(f"{base}/api/link/status", params={"key": lic})
        data = r.json() if r.status_code == 200 else {}
    except Exception as e:  # pragma: no cover - network/offline
        print(f"  (pairing bootstrap skipped: {type(e).__name__})")
        return
    if data.get("paired") and data.get("pairingSecret"):
        rec = _load_pairing()
        rec.update(
            {
                "pairing_secret": str(data["pairingSecret"]),
                "peer_product": data.get("peerProduct") or "coffer",
                "peer_identity": data.get("peerIdentity")
                or rec.get("peer_identity", ""),
                "peer_base_url": rec.get("peer_base_url") or COFFER_API_URL.rstrip("/"),
                "linked_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            }
        )
        _save_pairing(rec)
        print("  Coffer pairing restored from web (paired).")
    elif not _load_pairing().get("pairing_secret"):
        print("  No Coffer pairing linked to this account yet.")


def main() -> None:
    _bootstrap_profile()
    _bootstrap_pairing()  # restore Coffer pairing from web so a restart never comes up inert
    url = f"http://127.0.0.1:{PORT}"
    print(f"\n  AccountingQB local is live → {url}\n")
    if not os.environ.get("ACCOUNTINGQB_NO_OPEN"):
        try:
            if sys.platform == "darwin":
                subprocess.Popen(
                    ["open", url], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
                )
            elif sys.platform.startswith("win"):
                os.startfile(url)  # type: ignore[attr-defined]
            else:
                subprocess.Popen(
                    ["xdg-open", url],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
        except Exception:
            pass
    uvicorn.run(app, host="127.0.0.1", port=PORT, log_level="warning")


if __name__ == "__main__":
    main()
